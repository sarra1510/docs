"""
Plan de Charge QA/DEV - CACEIS Release 3.2.0
Génère un fichier Excel professionnel avec :
  - Suivi estimé / réel / restant (recalcul automatique)
  - Gantt mensuel (semaines) - statuts complets
  - Sections : DEV / QA / TNR / Fix Defect
  - Mise en forme couleurs + filtres + formules dynamiques
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.chart import BarChart, Reference

# ============================================================
# CONFIGURATION GLOBALE
# ============================================================
OUTPUT_FILE = "Plan_de_Charge_CACEIS_3_2_0.xlsx"

C_NAVY       = "002060"
C_BLUE_HDR   = "1F4E79"
C_BLUE_QA    = "BDD7EE"
C_GREEN_DEV  = "C6EFCE"
C_ORANGE_TNR = "FCE4D6"
C_YELLOW_FIX = "FFEB9C"
C_GREY_SCOPE = "D9D9D9"
C_WHITE      = "FFFFFF"
C_DARK_TEXT  = "1A1A1A"

STATUS_COLORS = {
    "Terminé":       ("E2EFDA", "375623"),
    "OK":            ("E2EFDA", "375623"),
    "En cours":      ("FFEB9C", "7D6608"),
    "Non commencé":  ("EDEDED", "595959"),
    "Planifié":      ("DDEBF7", "1F4E79"),
    "Bloquant":      ("F4CCCC", "9C0006"),
    "Bloqué":        ("F4CCCC", "9C0006"),
    "À faire":       ("FFF2CC", "7D6608"),
    "À venir":       ("DDEBF7", "1F4E79"),
    "Nouveau":       ("C9DAF8", "1F3864"),
    "Descoped":      ("D9D9D9", "595959"),
    "KO":            ("F4CCCC", "9C0006"),
    "In Progress":   ("FFEB9C", "7D6608"),
}

TYPE_COLORS = {
    "DEV":    C_GREEN_DEV,
    "QA":     C_BLUE_QA,
    "TNR":    C_ORANGE_TNR,
    "DEFECT": C_YELLOW_FIX,
    "SCOPE":  C_GREY_SCOPE,
}

thin = Side(border_style="thin",   color="B8B8B8")
med  = Side(border_style="medium", color="002060")

def border_thin():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def style_cell(cell, bg=None, fg=C_DARK_TEXT, bold=False, size=10,
               h="left", v="center", wrap=True, italic=False, border=True):
    if bg:
        cell.fill = fill(bg)
    cell.font = Font(bold=bold, color=fg, size=size, italic=italic, name="Calibri")
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border:
        cell.border = border_thin()

# ============================================================
# DONNÉES
# ============================================================
HEADER = [
    "N°", "Type", "Catégorie", "Tâche / Ticket", "Description",
    "Responsable", "Sprint", "Priorité", "Statut",
    "Estimé (j)", "Réel (j)", "Restant (j)", "% Avancement",
    "Date début", "Date fin prév.", "Commentaires"
]

DATA = [
    # ─── DEV - Release 3.2.0
    [1,  "DEV",    "Release",    "Doc Installation",          "Rédaction doc install : upgrade 2.1.21 → 3.0.0 → 3.2.0",   "Chakib",       "S18", "High",   "Terminé",      1.0, 1.0,  "", "", "2026-04-28", "2026-04-29", ""],
    [2,  "DEV",    "Release",    "Release Note",              "Revue et nettoyage de la release note",                      "Mehdi",        "S18", "Medium", "En cours",     1.0, 0.5,  "", "", "2026-04-29", "2026-04-30", ""],
    [3,  "DEV",    "Release",    "IMPRESS-27404",             "Multi scheduler - attente TEST-9968 COMET",                  "Amine",        "S19", "High",   "Bloquant",     2.0, 0.0,  "", "", "2026-05-04", "2026-05-05", "Décision sous 4h, impact release à confirmer"],
    # ─── DEV - Fix Defects TNR BLOCKER
    [4,  "DEFECT", "Fix Defect", "IMPRESS-27904",             "Ecart Lux (2) – Trouver root cause & fix",                  "Samir",        "S18", "Blocker","En cours",     2.0, 0.5,  "", "", "2026-04-29", "2026-05-01", "Raised by QA – Blocker TNR"],
    [5,  "DEFECT", "Fix Defect", "IMPRESS-27999",             "Ecart CH – root cause & fix (package only)",                "Samir",        "S18", "Blocker","À faire",      1.0, 0.0,  "", "", "2026-04-29", "2026-04-30", "ETA 29/04 – Raised by QA – Blocker TNR"],
    [6,  "DEFECT", "Fix Defect", "IMPRESS-28006",             "Ecart Lux – root cause à trouver puis fix",                 "",             "S19", "Blocker","En cours",     2.0, 0.5,  "", "", "2026-05-04", "2026-05-06", "Raised by QA – Blocker TNR"],
    [7,  "DEFECT", "Fix Defect", "IMPRESS-27995",             "Planification OGC – repro sur VM",                          "",             "S19", "High",   "En cours",     2.0, 0.5,  "", "", "2026-05-04", "2026-05-06", "Raised by PS – problème VM"],
    [8,  "DEFECT", "Fix Defect", "IMPRESS-27997",             "Audit Excel – tableau récap fin de fichier",                "",             "S19", "High",   "À faire",      1.5, 0.0,  "", "", "2026-05-04", "2026-05-05", "Raised by PS"],
    # ─── DEV - Descope
    [9,  "SCOPE",  "Descope",    "IMPRESS-27913",             "ERROR: missing report auditor validation → déjà en 2.1.21", "Chakib",       "S18", "Low",    "Descoped",     0.0, 0.0,  "", "", "",           "",           "Descope : déjà présent en 2.1.21"],
    [10, "SCOPE",  "Descope",    "IMPRESS-27230",             "FR – NIMI & Subsidiaries Templating title repetition",      "",             "S18", "Low",    "Descoped",     0.0, 0.0,  "", "", "",           "",           "Descope : déjà présent en 2.1.21"],
    # ─── QA - Setup
    [11, "QA",     "Setup",      "Package QA",                "Validation / retrait package Basma",                        "Basma",        "S18", "Medium", "En cours",     0.5, 0.25, "", "", "2026-04-29", "2026-04-30", "A retirer ?"],
    [12, "QA",     "Setup",      "Descope ECDF Workflows",    "Workflows ECDF (Imen Mess) – DESCOPE",                      "Imen M.",      "S18", "Low",    "Descoped",     1.0, 0.0,  "", "", "",           "",           "Descope décidé en réunion"],
    [13, "QA",     "Setup",      "Descope ECDF DQCs",         "DQCs ECDF (Rania) – DESCOPE",                               "Rania",        "S18", "Low",    "Descoped",     3.0, 0.0,  "", "", "",           "",           "Descope décidé en réunion"],
    [14, "QA",     "Setup",      "Descope TNR PRIIPS",        "TNRs PRIIPS packages – DESCOPE",                            "Rihab",        "S18", "Low",    "Descoped",     1.0, 0.0,  "", "", "",           "",           "Descope décidé en réunion"],
    # ─── QA - Validation Fiches
    [15, "QA",     "Validation", "IMPRESS-27949",             "DH7 : Incohérence ordre champs – Portfolio Valuation",      "Imen M.",      "S19", "High",   "Nouveau",      0.5, 0.0,  "", "", "2026-05-04", "2026-05-05", ""],
    [16, "QA",     "Validation", "IMPRESS-27913",             "[CACEIS] ERROR missing report auditor validation",          "",             "S19", "High",   "Nouveau",      0.5, 0.0,  "", "", "2026-05-04", "2026-05-05", ""],
    [17, "QA",     "Validation", "IMPRESS-27790",             "Convergence – Incompatible TEXT_TYPE list CACEIS/RefProd",  "",             "S19", "Medium", "Nouveau",      1.0, 0.0,  "", "", "2026-05-04", "2026-05-06", ""],
    [18, "QA",     "Validation", "IMPRESS-27819",             "Convergence RefProd – Bouton simulation Tag Label KO",      "",             "S19", "Medium", "Nouveau",      0.5, 0.0,  "", "", "2026-05-04", "2026-05-05", ""],
    [19, "QA",     "Validation", "IMPRESS-27997",             "Audit Excel – tableau récap fin de fichier",                "",             "S19", "High",   "Nouveau",      1.0, 0.0,  "", "", "2026-05-04", "2026-05-06", ""],
    [20, "QA",     "Validation", "IMPRESS-27995",             "[Caceis] Planifications OGC ne remontent pas sur Front",    "",             "S19", "High",   "Nouveau",      1.0, 0.0,  "", "", "2026-05-04", "2026-05-06", ""],
    [21, "QA",     "Validation", "IMPRESS-27999",             "[Caceis] Calculation discrepancies – Suisse report",        "",             "S18", "Blocker","Non commencé", 1.0, 0.0,  "", "", "2026-04-30", "2026-05-01", ""],
    [22, "QA",     "Validation", "IMPRESS-27904",             "[Caceis] Calculation discrepancies – Lux report (2)",       "",             "S18", "Blocker","Non commencé", 1.0, 0.0,  "", "", "2026-04-30", "2026-05-01", ""],
    [23, "QA",     "Validation", "IMPRESS-27775",             "Echec workflow Belges – compute timeout",                   "",             "S19", "Medium", "In Progress",  1.0, 0.5,  "", "", "2026-05-04", "2026-05-05", ""],
    [24, "QA",     "Validation", "IMPRESS-27850",             "Impossible de purger le rapport",                           "",             "S19", "Medium", "In Progress",  1.0, 0.5,  "", "", "2026-05-04", "2026-05-05", ""],
    [25, "QA",     "Validation", "IMPRESS-27950",             "[CACEIS] Error: audit data download failed",                "",             "S19", "High",   "In Progress",  1.0, 0.5,  "", "", "2026-05-04", "2026-05-05", ""],
    # ─── TNR - Scénarios Caceis
    [26, "TNR",    "TNR FR",     "Scénarios FR",              "Finalisation scénarios France CACEIS",                      "",             "S19", "High",   "En cours",     2.0, 0.5,  "", "", "2026-05-04", "2026-05-07", "2j estimé"],
    [27, "TNR",    "TNR LUX",   "608786 – Comparaison KO",   "Comparaison des rapports KO (DHRD-122486)",                 "",             "S19", "Blocker","KO",           1.0, 0.5,  "", "", "2026-05-04", "2026-05-06", "DHRD-122486"],
    [28, "TNR",    "TNR LUX",   "999140",                    "Scénario LUX 999140",                                        "",             "S19", "Medium", "Terminé",      0.5, 0.5,  "", "", "2026-05-04", "2026-05-05", "OK"],
    [29, "TNR",    "TNR LUX",   "207597 (x2)",               "Scénarios LUX 207597 – TO DO",                              "",             "S19", "High",   "À faire",      1.0, 0.0,  "", "", "2026-05-05", "2026-05-06", "TO DO"],
    [30, "TNR",    "TNR LUX",   "999496 / 170399 / 246841",  "Scénarios LUX – TO DO",                                     "",             "S19", "High",   "À faire",      1.5, 0.0,  "", "", "2026-05-06", "2026-05-08", "TO DO"],
    [31, "TNR",    "TNR LUX",   "448519 / 517192",           "Scénarios LUX – TO DO",                                     "",             "S19", "High",   "À faire",      1.0, 0.0,  "", "", "2026-05-07", "2026-05-08", "TO DO"],
    [32, "TNR",    "TNR CH",    "357904 / 637925",           "Scénarios CH – OK",                                         "",             "S19", "Medium", "Terminé",      0.5, 0.5,  "", "", "2026-05-04", "2026-05-04", "OK"],
    [33, "TNR",    "TNR CH",    "999328",                    "Rapport vide (same as ref)",                                 "",             "S19", "Medium", "En cours",     0.5, 0.25, "", "", "2026-05-04", "2026-05-05", "Rapport vide = ref ?"],
    [34, "TNR",    "TNR CH",    "999856",                    "Scénario KO (IMPRESS-27999)",                               "",             "S19", "Blocker","KO",           1.0, 0.5,  "", "", "2026-05-04", "2026-05-06", "Lié IMPRESS-27999"],
    [35, "TNR",    "TNR CH",    "234312",                    "Scénario CH 234312",                                        "",             "S19", "High",   "In Progress",  1.0, 0.5,  "", "", "2026-05-05", "2026-05-06", ""],
    [36, "TNR",    "TNR CH",    "377138",                    "Use case métier à voir avec PS",                            "",             "S20", "Medium", "À faire",      1.0, 0.0,  "", "", "2026-05-11", "2026-05-12", "Attente PS"],
    [37, "TNR",    "TNR BE",    "Scénarios BE",              "Finalisation scénarios Belgique",                           "",             "S20", "High",   "À faire",      1.0, 0.0,  "", "", "2026-05-11", "2026-05-12", "1j estimé"],
    [38, "TNR",    "TNR CACEIS","TNRs CACEIS full",          "TNRs CACEIS (les 2 Imen) – 2j",                             "Imen M./Imen", "S20", "High",   "Planifié",     2.0, 0.0,  "", "", "2026-05-11", "2026-05-14", "7 md full QA besoin si build complet"],
]

# ============================================================
# WORKBOOK
# ============================================================
wb = openpyxl.Workbook()

# ────────────────────────────────────────────────────────────
# ONGLET 1 : PLAN DE CHARGE
# ────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Plan de Charge"

# Titre principal
ws.merge_cells("A1:P1")
title_cell = ws["A1"]
title_cell.value = "PLAN DE CHARGE QA / DEV — CACEIS Release 3.2.0"
style_cell(title_cell, bg=C_NAVY, fg=C_WHITE, bold=True, size=14, h="center", border=False)
ws.row_dimensions[1].height = 32

# Sous-titre
ws.merge_cells("A2:P2")
sub = ws["A2"]
sub.value = "Période : Mai – Juin 2026  |  Équipe : QA CACEIS + DEV IMPRESS  |  MAJ : 02/05/2026"
style_cell(sub, bg=C_BLUE_HDR, fg=C_WHITE, bold=False, size=10, h="center", italic=True, border=False)
ws.row_dimensions[2].height = 18

# En-têtes colonnes
for col_idx, col_name in enumerate(HEADER, 1):
    cell = ws.cell(row=3, column=col_idx, value=col_name)
    style_cell(cell, bg=C_BLUE_HDR, fg=C_WHITE, bold=True, size=10, h="center", border=True)
ws.row_dimensions[3].height = 28

# Données
DATA_START_ROW = 4
for row_idx, row_data in enumerate(DATA, DATA_START_ROW):
    n, typ, cat, tache, desc, resp, sprint, prio, statut, estime, reel = row_data[:11]
    date_d, date_f, comment = row_data[13], row_data[14], row_data[15]
    ws.row_dimensions[row_idx].height = 20

    for c_i, val in enumerate([n, typ, cat, tache, desc, resp, sprint, prio, statut, estime, reel], 1):
        cell = ws.cell(row=row_idx, column=c_i, value=val)
        bg = TYPE_COLORS.get(typ, C_WHITE) if c_i <= 9 else C_WHITE
        style_cell(cell, bg=bg, size=9, h="center" if c_i in [1, 6, 7, 8, 10, 11] else "left")

    # Restant = formule
    restant_cell = ws.cell(row=row_idx, column=12)
    restant_cell.value = f"=MAX(0,J{row_idx}-K{row_idx})"
    style_cell(restant_cell, bg="FFF2CC", size=9, h="center")

    # % Avancement = formule
    pct_cell = ws.cell(row=row_idx, column=13)
    pct_cell.value = f'=IF(J{row_idx}=0,"N/A",MIN(1,K{row_idx}/J{row_idx}))'
    pct_cell.number_format = "0%"
    style_cell(pct_cell, bg="DDEBF7", size=9, h="center")

    # Dates
    ws.cell(row=row_idx, column=14, value=date_d)
    style_cell(ws.cell(row=row_idx, column=14), size=9, h="center")
    ws.cell(row=row_idx, column=15, value=date_f)
    style_cell(ws.cell(row=row_idx, column=15), size=9, h="center")

    # Commentaire
    ws.cell(row=row_idx, column=16, value=comment)
    style_cell(ws.cell(row=row_idx, column=16), size=9, h="left")

    # Couleur statut
    statut_cell = ws.cell(row=row_idx, column=9)
    sc = STATUS_COLORS.get(statut, ("FFFFFF", C_DARK_TEXT))
    statut_cell.fill = fill(sc[0])
    statut_cell.font = Font(bold=True, color=sc[1], size=9, name="Calibri")

    # Couleur priorité
    prio_colors = {
        "Blocker": ("F4CCCC", "9C0006"),
        "High":    ("FCE4D6", "C55A11"),
        "Medium":  ("FFF2CC", "7D6608"),
        "Low":     ("EDEDED", "595959"),
    }
    pc = prio_colors.get(prio, ("FFFFFF", C_DARK_TEXT))
    prio_cell = ws.cell(row=row_idx, column=8)
    prio_cell.fill = fill(pc[0])
    prio_cell.font = Font(bold=True, color=pc[1], size=9, name="Calibri")

# Ligne TOTAL
TOTAL_ROW = DATA_START_ROW + len(DATA)
ws.row_dimensions[TOTAL_ROW].height = 22
ws.merge_cells(f"A{TOTAL_ROW}:I{TOTAL_ROW}")
total_label = ws[f"A{TOTAL_ROW}"]
total_label.value = "TOTAL"
style_cell(total_label, bg=C_NAVY, fg=C_WHITE, bold=True, size=10, h="center")

for c_i, col_letter in enumerate(["J", "K", "L"], 10):
    cell = ws.cell(row=TOTAL_ROW, column=c_i)
    cell.value = f"=SUM({col_letter}{DATA_START_ROW}:{col_letter}{TOTAL_ROW - 1})"
    style_cell(cell, bg=C_NAVY, fg=C_WHITE, bold=True, size=10, h="center")

pct_total = ws.cell(row=TOTAL_ROW, column=13)
pct_total.value = f"=IF(J{TOTAL_ROW}=0,0,K{TOTAL_ROW}/J{TOTAL_ROW})"
pct_total.number_format = "0%"
style_cell(pct_total, bg=C_NAVY, fg=C_WHITE, bold=True, size=10, h="center")
for c_i in [14, 15, 16]:
    ws.cell(row=TOTAL_ROW, column=c_i).fill = fill(C_NAVY)

# Largeurs colonnes
col_widths = [5, 8, 14, 20, 42, 12, 8, 10, 15, 10, 8, 10, 12, 13, 13, 35]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADER))}3"
ws.freeze_panes = "A4"

# DataBar Restant & % Avancement
last_data_row = TOTAL_ROW - 1
ws.conditional_formatting.add(
    f"L{DATA_START_ROW}:L{last_data_row}",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10, color="638EC6")
)
ws.conditional_formatting.add(
    f"M{DATA_START_ROW}:M{last_data_row}",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="70AD47")
)

# ────────────────────────────────────────────────────────────
# ONGLET 2 : SUIVI HEBDOMADAIRE (Gantt)
# ────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Suivi Hebdomadaire")

WEEKS = [
    "S18\n28/04–02/05",
    "S19\n05/05–09/05",
    "S20\n12/05–16/05",
    "S21\n19/05–23/05",
    "S22\n26/05–30/05",
]

ws2.merge_cells("A1:M1")
ws2["A1"].value = "SUIVI HEBDOMADAIRE — CACEIS Release 3.2.0"
style_cell(ws2["A1"], bg=C_NAVY, fg=C_WHITE, bold=True, size=13, h="center", border=False)
ws2.row_dimensions[1].height = 30

GANTT_HEADERS = [
    "N°", "Type", "Tâche / Ticket", "Responsable", "Statut",
    "Estimé (j)", "Réel (j)", "Restant (j)"
] + WEEKS

for ci, h in enumerate(GANTT_HEADERS, 1):
    cell = ws2.cell(row=2, column=ci, value=h)
    style_cell(cell, bg=C_BLUE_HDR, fg=C_WHITE, bold=True, size=9, h="center")
ws2.row_dimensions[2].height = 40

SPRINT_WEEK_MAP = {"S18": 0, "S19": 1, "S20": 2, "S21": 3, "S22": 4}
GANTT_COLORS = {
    "DEV":    "70AD47",
    "QA":     "5B9BD5",
    "TNR":    "ED7D31",
    "DEFECT": "FFC000",
    "SCOPE":  "A5A5A5",
}

for ri, row_data in enumerate(DATA, 3):
    n, typ, cat, tache, _, resp, sprint, _, statut, estime, reel = row_data[:11]
    ws2.row_dimensions[ri].height = 28

    for ci, val in enumerate([n, typ, tache, resp, statut, estime, reel], 1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        bg = TYPE_COLORS.get(typ, C_WHITE) if ci <= 2 else C_WHITE
        style_cell(cell, bg=bg, size=9, h="center" if ci in [1, 6, 7] else "left")

    # Restant formule
    rest = ws2.cell(row=ri, column=8)
    rest.value = f"=MAX(0,F{ri}-G{ri})"
    style_cell(rest, bg="FFF2CC", size=9, h="center")

    # Colorier la semaine du sprint dans le Gantt — STATUT COMPLET (pas tronqué)
    week_idx = SPRINT_WEEK_MAP.get(sprint, -1)
    for wc in range(9, 14):
        ec = ws2.cell(row=ri, column=wc)
        if wc == (9 + week_idx) and week_idx >= 0:
            gantt_val = "—" if statut == "Descoped" else statut
            ec.value = gantt_val
            ec.fill = fill(GANTT_COLORS.get(typ, "CCCCCC"))
            ec.font = Font(bold=True, color=C_WHITE, size=8, name="Calibri")
            ec.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            ec.fill = fill("F5F5F5")
        ec.border = border_thin()

# Largeurs onglet 2 — colonnes Gantt plus larges pour afficher statut complet
ws2_widths = [5, 8, 22, 12, 15, 9, 8, 9, 16, 16, 16, 16, 16]
for i, w in enumerate(ws2_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = "A3"
ws2.auto_filter.ref = f"A2:{get_column_letter(len(GANTT_HEADERS))}2"

# ────────────────────────────────────────────────────────────
# ONGLET 3 : RÉCAP & KPIs
# ────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Récap & KPIs")
ws3.merge_cells("A1:F1")
ws3["A1"].value = "RÉCAPITULATIF PAR TYPE — CACEIS Release 3.2.0"
style_cell(ws3["A1"], bg=C_NAVY, fg=C_WHITE, bold=True, size=13, h="center", border=False)
ws3.row_dimensions[1].height = 30

recap_headers = ["Type", "Nb Tâches", "Estimé Total (j)", "Réel Total (j)", "Restant (j)", "% Avancement"]
for ci, h in enumerate(recap_headers, 1):
    cell = ws3.cell(row=2, column=ci, value=h)
    style_cell(cell, bg=C_BLUE_HDR, fg=C_WHITE, bold=True, size=10, h="center")

pdc_sheet = "Plan de Charge"
types = ["DEV", "QA", "TNR", "DEFECT", "SCOPE"]
for ri, typ in enumerate(types, 3):
    ws3.cell(row=ri, column=1, value=typ)
    style_cell(ws3.cell(row=ri, column=1), bg=TYPE_COLORS.get(typ, C_WHITE), bold=True, size=10, h="center")
    ws3.cell(row=ri, column=2).value = f"=COUNTIF('{pdc_sheet}'!B4:B{TOTAL_ROW-1},A{ri})"
    ws3.cell(row=ri, column=3).value = f"=SUMIF('{pdc_sheet}'!B4:B{TOTAL_ROW-1},A{ri},'{pdc_sheet}'!J4:J{TOTAL_ROW-1})"
    ws3.cell(row=ri, column=4).value = f"=SUMIF('{pdc_sheet}'!B4:B{TOTAL_ROW-1},A{ri},'{pdc_sheet}'!K4:K{TOTAL_ROW-1})"
    ws3.cell(row=ri, column=5).value = f"=SUMIF('{pdc_sheet}'!B4:B{TOTAL_ROW-1},A{ri},'{pdc_sheet}'!L4:L{TOTAL_ROW-1})"
    pct = ws3.cell(row=ri, column=6)
    pct.value = f"=IF(C{ri}=0,0,D{ri}/C{ri})"
    pct.number_format = "0%"
    for ci in range(2, 7):
        style_cell(ws3.cell(row=ri, column=ci), size=10, h="center")

# Totaux
tr = 3 + len(types)
ws3.cell(row=tr, column=1, value="TOTAL")
style_cell(ws3.cell(row=tr, column=1), bg=C_NAVY, fg=C_WHITE, bold=True, h="center")
for ci in range(2, 7):
    cell = ws3.cell(row=tr, column=ci)
    col_letter = get_column_letter(ci)
    if ci < 6:
        cell.value = f"=SUM({col_letter}3:{col_letter}{tr-1})"
    else:
        cell.value = f"=IF(C{tr}=0,0,D{tr}/C{tr})"
        cell.number_format = "0%"
    style_cell(cell, bg=C_NAVY, fg=C_WHITE, bold=True, h="center")

for i, w in enumerate([14, 12, 18, 14, 12, 16], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A3"

# Graphique Estimé vs Réel
chart = BarChart()
chart.type = "col"
chart.grouping = "clustered"
chart.title = "Estimé vs Réel par Type"
chart.style = 10
chart.y_axis.title = "Jours"
chart.x_axis.title = "Type"
data_ref = Reference(ws3, min_col=3, max_col=4, min_row=2, max_row=2 + len(types))
cats_ref = Reference(ws3, min_col=1, min_row=3, max_row=2 + len(types))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws3.add_chart(chart, "A10")

# ============================================================
# SAUVEGARDE
# ============================================================
wb.save(OUTPUT_FILE)
print(f"✅ Fichier '{OUTPUT_FILE}' généré avec succès !")
print(f"   → 3 onglets : 'Plan de Charge' | 'Suivi Hebdomadaire' | 'Récap & KPIs'")
print(f"   → {len(DATA)} tâches | Formules dynamiques | Filtres | Graphiques")
print(f"   → Statuts affichés complets dans le Gantt (correction tronquage)")
print(f"\n   Pour ouvrir automatiquement (Windows) :")
print(f"   import os; os.startfile('{OUTPUT_FILE}')")
